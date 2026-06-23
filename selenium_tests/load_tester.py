import os
import sys
import time
import json
import threading
import requests
import openpyxl
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Adjust path to import config
BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

import config

# Configuration
CONCURRENT_USERS = 20
TEST_DURATION_SECS = 30
API_URL = "http://127.0.0.1:8001"
HEADERS = {"Authorization": f"Bearer {config.NEXTAUTH_SECRET}"}

# Thread-safe lists to collect statistics
response_times_ms = []
success_count = 0
failure_count = 0
detailed_results = []
stats_lock = threading.Lock()

def check_server_online():
    try:
        r = requests.get(f"{API_URL}/health", timeout=3)
        return r.status_code == 200
    except Exception:
        return False

def virtual_user_thread(stop_event):
    global success_count, failure_count
    
    session = requests.Session()
    session.headers.update(HEADERS)
    
    request_counter = 0
    
    while not stop_event.is_set():
        with stats_lock:
            if len(detailed_results) >= 500:
                stop_event.set()
                break
                
        endpoint = "/health" if request_counter % 2 == 0 else "/search"
        url = f"{API_URL}{endpoint}"
        
        t0 = time.time()
        try:
            if endpoint == "/health":
                r = session.get(url, timeout=5)
            else:
                r = session.post(url, json={"query": "performance test query", "top_k": 3}, timeout=5)
                
            elapsed_ms = (time.time() - t0) * 1000
            
            with stats_lock:
                if len(detailed_results) >= 500:
                    stop_event.set()
                    break
                    
                status_code = r.status_code
                if r.status_code in (200, 201):
                    success_count += 1
                    status_text = "PASS"
                else:
                    failure_count += 1
                    status_text = "FAIL"
                    
                response_times_ms.append(elapsed_ms)
                detailed_results.append({
                    "id": f"TC_LOAD_{len(detailed_results)+1:03d}",
                    "endpoint": endpoint,
                    "response_time": elapsed_ms,
                    "status_code": status_code,
                    "status": status_text
                })
        except Exception:
            elapsed_ms = (time.time() - t0) * 1000
            with stats_lock:
                if len(detailed_results) >= 500:
                    stop_event.set()
                    break
                failure_count += 1
                response_times_ms.append(elapsed_ms)
                detailed_results.append({
                    "id": f"TC_LOAD_{len(detailed_results)+1:03d}",
                    "endpoint": endpoint,
                    "response_time": elapsed_ms,
                    "status_code": 0,
                    "status": "ERROR"
                })
                
        request_counter += 1
        time.sleep(0.01)

def run_load_test():
    print("=" * 60)
    print("PROMPT PILOT CONCURRENT LOAD TESTER (500 TESTS)")
    print("=" * 60)
    print(f"Target Server: {API_URL}")
    print(f"Concurrent Virtual Users: {CONCURRENT_USERS}")
    
    if not check_server_online():
        print(f"\n[Warning] Context server at {API_URL} is not responding.")
        print("Starting test anyway (metrics will reflect connection failures)...")
    else:
        print("\n[Online] Context server is active. Beginning load run...")
        
    stop_event = threading.Event()
    threads = []
    
    for i in range(CONCURRENT_USERS):
        t = threading.Thread(target=virtual_user_thread, args=(stop_event,))
        t.daemon = True
        threads.append(t)
        t.start()
        
    print(f"Spawned {CONCURRENT_USERS} concurrent workers. Collecting 500 requests...")
    
    start_time = time.time()
    last_print = start_time
    
    # Wait until we reach 500 requests or timeout
    while not stop_event.is_set():
        with stats_lock:
            curr_count = len(detailed_results)
            if curr_count >= 500:
                stop_event.set()
                break
        time.sleep(0.5)
        now = time.time()
        if now - last_print >= 3:
            with stats_lock:
                tot = len(detailed_results)
                elapsed = now - start_time
                curr_rps = tot / elapsed if elapsed > 0 else 0
                print(f"  [Progress] {int(elapsed)}s elapsed | Total Requests: {tot}/500 | RPS: {curr_rps:.1f}")
            last_print = now
            
        if now - start_time >= TEST_DURATION_SECS:
            print("[Timeout] Target execution time reached before 500 requests.")
            stop_event.set()
            break
            
    # Stop remaining threads
    stop_event.set()
    time.sleep(1)
    
    total_elapsed = time.time() - start_time
    
    with stats_lock:
        tot_requests = len(detailed_results)
        succ = success_count
        fail = failure_count
        resp_times = list(response_times_ms)
        results_list = list(detailed_results)
        
    avg_rps = tot_requests / total_elapsed if total_elapsed > 0 else 0
    
    if resp_times:
        min_time = min(resp_times)
        max_time = max(resp_times)
        avg_time = sum(resp_times) / len(resp_times)
    else:
        min_time = 0
        max_time = 0
        avg_time = 0
        
    print("\n" + "=" * 60)
    print("PERFORMANCE RESULTS SUMMARY")
    print("=" * 60)
    print(f"Total Requests Sent:   {tot_requests}")
    print(f"Successful Requests:   {succ}")
    print(f"Failed Requests:       {fail}")
    print(f"Success Rate:          {(succ/tot_requests*100) if tot_requests else 0:.1f}%")
    print(f"Average RPS:           {avg_rps:.1f} req/sec")
    print(f"Response Times:")
    print(f"  Min:     {min_time:.1f}ms")
    print(f"  Max:     {max_time:.1f}ms")
    print(f"  Average: {avg_time:.1f}ms")
    print("=" * 60)
    
    generate_excel_report(tot_requests, succ, fail, avg_rps, min_time, max_time, avg_time, results_list)
    generate_markdown_report(tot_requests, succ, fail, avg_rps, min_time, max_time, avg_time)

def generate_excel_report(total, success, failure, rps, min_time, max_time, avg_time, results_list):
    report_path = BASE_DIR / "reports" / "performance_load_report.xlsx"
    report_path.parent.mkdir(exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Summary"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1F4E79")
    bold_font = Font(name=font_family, size=10, bold=True)
    regular_font = Font(name=font_family, size=10)
    
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    metric_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    info_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Title Block
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "PROMPT PILOT — BASELINE LOAD PERFORMANCE REPORT"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Metadata
    ws.merge_cells("A2:E2")
    meta_cell = ws["A2"]
    meta_cell.value = f"Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: {API_URL}"
    meta_cell.font = Font(name=font_family, size=9, italic=True, color="595959")
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Section 1: Parameters
    ws.cell(row=4, column=1, value="1. Test Execution Settings").font = section_font
    
    params = [
        ("Concurrent Virtual Users", f"{CONCURRENT_USERS} Users"),
        ("Request Protocols", "GET /health & POST /search"),
        ("Target Host API Port", f"{API_URL} (:8001)")
    ]
    
    current_row = 5
    for name, value in params:
        c1 = ws.cell(row=current_row, column=1, value=name)
        c2 = ws.cell(row=current_row, column=2, value=value)
        c1.font = bold_font
        c2.font = regular_font
        c1.border = thin_border
        c2.border = thin_border
        current_row += 1
        
    # Section 2: Core Summary Metrics
    ws.cell(row=9, column=1, value="2. Performance Load Results").font = section_font
    
    headers = ["Metric Category", "Description", "Target Threshold", "Measured Value", "Status"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[10].height = 25
        
    success_rate = (success / total * 100) if total else 0
    
    results_metrics = [
        ("Total Requests", "Total HTTP payloads dispatched", "500", total, "PASS" if total >= 500 else "FAIL"),
        ("Success Count", "Requests returning HTTP 200/201", "N/A", success, "INFO"),
        ("Failure Count", "Requests returning exception or errors", "0", failure, "PASS" if failure == 0 else "FAIL"),
        ("Success Rate", "Percentage of successful payloads", ">= 95.0 %", success_rate, "PASS" if success_rate >= 95.0 else "FAIL"),
        ("Average RPS", "Requests processed per second", ">= 80 req/sec", rps, "PASS" if rps >= 80 else "WARNING"),
        ("Min Response Time", "Fastest response roundtrip", "N/A", min_time, "INFO"),
        ("Max Response Time", "Slowest response roundtrip", "< 2000 ms", max_time, "PASS" if max_time < 2000 else "WARNING"),
        ("Average Response Time", "Average latency overall", "< 1000 ms", avg_time, "PASS" if avg_time < 1000 else "WARNING")
    ]
    
    current_row = 11
    for cat, desc, threshold, val, status in results_metrics:
        ws.cell(row=current_row, column=1, value=cat).font = bold_font
        ws.cell(row=current_row, column=2, value=desc).font = regular_font
        ws.cell(row=current_row, column=3, value=threshold).font = regular_font
        
        val_cell = ws.cell(row=current_row, column=4, value=val)
        val_cell.font = bold_font
        val_cell.alignment = Alignment(horizontal="right")
        
        if cat == "Success Rate":
            val_cell.number_format = '0.0'
        elif isinstance(val, float):
            val_cell.number_format = '0.0'
            
        status_cell = ws.cell(row=current_row, column=5, value=status)
        status_cell.font = bold_font
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if status == "PASS":
            status_cell.fill = pass_fill
        elif status == "FAIL":
            status_cell.fill = fail_fill
        elif status == "WARNING":
            status_cell.fill = warning_fill
        else:
            status_cell.fill = info_fill
            
        for col in range(1, 6):
            c = ws.cell(row=current_row, column=col)
            c.border = thin_border
            if col != 5:
                c.fill = metric_fill
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Section 3: Detailed Load Test Cases (500 Rows)
    current_row += 1
    ws.cell(row=current_row, column=1, value="3. Detailed Load Test Cases (500 Requests)").font = section_font
    current_row += 1
    
    detail_header_row = current_row
    detail_headers = ["Test Case ID", "Target Endpoint", "Response Time (ms)", "Status Code", "Status"]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    for res in results_list:
        ws.cell(row=current_row, column=1, value=res["id"]).font = bold_font
        ws.cell(row=current_row, column=2, value=res["endpoint"]).font = regular_font
        
        rt_cell = ws.cell(row=current_row, column=3, value=res["response_time"])
        rt_cell.font = regular_font
        rt_cell.number_format = '0.0'
        rt_cell.alignment = Alignment(horizontal="right")
        
        sc_cell = ws.cell(row=current_row, column=4, value=res["status_code"])
        sc_cell.font = regular_font
        sc_cell.alignment = Alignment(horizontal="center")
        
        status_cell = ws.cell(row=current_row, column=5, value=res["status"])
        status_cell.font = bold_font
        status_cell.alignment = Alignment(horizontal="center")
        if res["status"] in ("PASS", "SECURE"):
            status_cell.fill = pass_fill
        else:
            status_cell.fill = fail_fill
            
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
            if col != 5:
                ws.cell(row=current_row, column=col).fill = metric_fill
                
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        
    # Enable Autofilter dropdowns on detailed headers row
    ws.auto_filter.ref = f"A{detail_header_row}:E{current_row - 1}"
        
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    try:
        wb.save(str(report_path))
        print(f"[Report] Excel performance report saved successfully at: {report_path}")
    except Exception as e:
        print(f"[Error] Failed to save Excel report: {e}")

def generate_markdown_report(total, success, failure, rps, min_time, max_time, avg_time):
    summary_path = BASE_DIR.parent / "PERFORMANCE_TEST_SUMMARY.md"
    verdict = "STABLE" if failure == 0 and avg_time < 350 and rps >= 80 else "DEGRADED"
    
    content = f"""# PromptPilot Performance & Load Test Summary

**Server Status:** {verdict}
**Concurrency Target:** 500 total requests processed by concurrent workers

## Load Test Metrics

| Metric | Measured Value | Target Threshold | Status |
|--------|----------------|------------------|--------|
| **Total Requests Sent** | {total} | 500 | {'✅ PASS' if total >= 500 else '❌ FAIL'} |
| **Successful Requests** | {success} | N/A | info |
| **Failed Requests** | {failure} | 0 | {'✅ PASS' if failure == 0 else '❌ FAIL'} |
| **Success Rate** | {(success/total*100) if total else 0:.1f}% | 100.0% | {'✅ PASS' if failure == 0 else '❌ FAIL'} |
| **Average RPS** | {rps:.1f} req/sec | >= 80 req/sec | {'✅ PASS' if rps >= 80 else '⚠️ WARNING'} |
| **Min Response Time** | {min_time:.1f} ms | N/A | info |
| **Max Response Time** | {max_time:.1f} ms | < 2000 ms | {'✅ PASS' if max_time < 2000 else '⚠️ SLOW'} |
| **Average Response Time** | {avg_time:.1f} ms | < 300 ms | {'✅ PASS' if avg_time < 300 else '⚠️ SLOW'} |

## Test Parameters
- **Concurrent Users:** {CONCURRENT_USERS} threads
- **Request Target:** exactly 500 requests total
- **Endpoints Targeted:** `GET /health` and `POST /search`
- **Execution Timestamp:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Artifact Outputs
- Excel spreadsheet: `selenium_tests/reports/performance_load_report.xlsx`
"""
    try:
        summary_path.write_text(content, encoding="utf-8")
        print(f"[Report] Markdown summary report saved successfully at: {summary_path}")
    except Exception as e:
        print(f"[Error] Failed to write Markdown summary: {e}")

if __name__ == "__main__":
    run_load_test()
