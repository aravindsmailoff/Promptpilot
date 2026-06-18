import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class SecurityReporter:
    """Handles compilation of vulnerability scan results and exporting a premium Excel report."""
    
    def __init__(self):
        self.results = []
        
    def add_result(self, test_id: str, category: str, description: str, payload: str, expected: str, actual: str, risk: str, status: str, remediation: str, elapsed_time: float):
        """Append a security test execution record."""
        self.results.append({
            "id": test_id,
            "category": category,
            "description": description,
            "payload": payload,
            "expected": expected,
            "actual": actual,
            "risk": risk,
            "status": status.upper(),
            "remediation": remediation,
            "time": round(elapsed_time, 2),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        print(f"[{status.upper()}] {test_id} - {category}: {description} ({round(elapsed_time, 2)}s)")

    def generate_excel_report(self, report_path: Path, target_url: str):
        """Write all logged results to a highly stylized Excel sheet."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Security Scan Summary"
        
        # Show gridlines
        ws.views.sheetView[0].showGridLines = True
        
        # Typography & Colors
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
        header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10, color="000000")
        bold_data_font = Font(name=font_family, size=10, bold=True)
        italic_meta_font = Font(name=font_family, size=9, italic=True, color="595959")
        
        title_fill = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid") # Dark Charcoal
        header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid") # Dark Gray
        
        # Soft status fills
        secure_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green (Pass/Secure)
        vulnerable_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red (Fail/Vulnerable)
        
        # Soft risk level fills
        high_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        med_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        low_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Title Block
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = "PROMPTPILOT — VULNERABILITY TESTING AND DAST REPORT"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Metadata
        ws.merge_cells("A2:J2")
        meta_cell = ws["A2"]
        meta_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: {target_url}"
        meta_cell.font = italic_meta_font
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        
        # Spacer
        ws.row_dimensions[3].height = 10
        
        # Table Headers
        headers = [
            "Test ID", "Category", "Description", "Test Payload / Vector", 
            "Expected Outcome", "Actual Outcome", "Risk Level", 
            "Status", "Remediation & Security Control", "Duration (s)"
        ]
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[4].height = 25
        
        current_row = 5
        if not self.results:
            ws.cell(row=current_row, column=1, value="WARNING").font = bold_data_font
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
            warn_cell = ws.cell(row=current_row, column=2, value="NO TEST RESULTS CAPTURED — check conftest hooks and run_test() logging")
            warn_cell.font = data_font
            warn_cell.fill = vulnerable_fill
            current_row += 1

        for res in self.results:
            ws.cell(row=current_row, column=1, value=res["id"]).font = bold_data_font
            ws.cell(row=current_row, column=2, value=res["category"]).font = bold_data_font
            ws.cell(row=current_row, column=3, value=res["description"]).font = data_font
            ws.cell(row=current_row, column=4, value=res["payload"]).font = data_font
            ws.cell(row=current_row, column=5, value=res["expected"]).font = data_font
            ws.cell(row=current_row, column=6, value=res["actual"]).font = data_font
            
            # Risk formatting
            risk_cell = ws.cell(row=current_row, column=7, value=res["risk"])
            risk_cell.font = bold_data_font
            risk_cell.alignment = Alignment(horizontal="center", vertical="center")
            if res["risk"].lower() == "high":
                risk_cell.fill = high_fill
            elif res["risk"].lower() == "medium":
                risk_cell.fill = med_fill
            else:
                risk_cell.fill = low_fill
                
            # Status formatting
            status_cell = ws.cell(row=current_row, column=8, value=res["status"])
            status_cell.font = bold_data_font
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            if res["status"] in ("SECURE", "PASS"):
                status_cell.fill = secure_fill
            else:
                status_cell.fill = vulnerable_fill
                
            ws.cell(row=current_row, column=9, value=res["remediation"]).font = data_font
            
            dur_cell = ws.cell(row=current_row, column=10, value=res["time"])
            dur_cell.font = data_font
            dur_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Apply layout rules
            for col in range(1, 11):
                c = ws.cell(row=current_row, column=col)
                c.border = thin_border
                if col not in (7, 8, 10):
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                    
            ws.row_dimensions[current_row].height = 55
            current_row += 1
            
        # Fit columns to text width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 3 and cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 30), 12)
            
        # Smart save: if file is locked (open in Excel), save to a timestamped fallback
        report_path = Path(report_path)
        save_path = report_path.resolve()
        try:
            wb.save(str(save_path))
            print(f"[SecurityReporter] Report saved: {save_path}")
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            save_path = report_path.parent / f"vulnerability_report_{timestamp}.xlsx"
            wb.save(str(save_path.resolve()))
            print(f"[SecurityReporter] Primary file locked — close Excel and re-run. Saved to: {save_path.resolve()}")
        data_rows = len(self.results)
        print(f"[SecurityReporter] Total entries: {data_rows} data row(s) written")
        if data_rows == 0:
            print("[SecurityReporter] WARNING: Report has no test data rows (headers only).")
        return save_path
