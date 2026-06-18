import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class TestReporter:
    """Handles compilation of test results, taking screenshots, and exporting a stylized Excel report for Appium."""
    
    def __init__(self):
        self.results = []
        
    def add_result(self, test_id: str, component: str, description: str, steps: list, expected: str, actual: str, status: str, elapsed_time: float, screenshot_path: str = None):
        """Append a test execution record."""
        steps_str = "\n".join([f"{i+1}. {step}" for i, step in enumerate(steps)])
        self.results.append({
            "id": test_id,
            "component": component,
            "description": description,
            "steps": steps_str,
            "expected": expected,
            "actual": actual,
            "status": status.upper(),
            "time": round(elapsed_time, 2),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "screenshot": screenshot_path
        })
        print(f"[{status.upper()}] {test_id} - {description} ({round(elapsed_time, 2)}s)")

    def take_screenshot(self, driver, test_id: str) -> str:
        """Capture screenshot and return the relative path from the reports directory."""
        from config import SCREENSHOTS_DIR
        filename = f"{test_id}_{int(time.time())}.png"
        full_path = SCREENSHOTS_DIR / filename
        try:
            driver.save_screenshot(str(full_path))
            # Return path relative to the Excel file so it's clickable
            return f"screenshots/{filename}"
        except Exception as e:
            print(f"[TestReporter] Failed to save screenshot: {e}")
            return None

    def generate_excel_report(self):
        """Write all logged results to a stylized Excel sheet."""
        from config import REPORT_FILE_PATH
        
        # Create a new workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Execution Summary"
        
        # Show gridlines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Colors & Styling
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10, color="000000")
        bold_data_font = Font(name=font_family, size=10, bold=True)
        link_font = Font(name=font_family, size=10, color="0563C1", underline="single")
        
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Navy Blue
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Medium Blue
        
        # Soft fills for statuses
        pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
        fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red/orange
        error_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Write Title Block
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = "PROMPT PILOT — APPIUM MOBILE AUTOMATION TEST REPORT"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Add metadata row
        ws.merge_cells("A2:J2")
        meta_cell = ws["A2"]
        meta_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target Application: {config_target_url()}"
        meta_cell.font = Font(name=font_family, size=10, italic=True, color="595959")
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        
        # Empty row
        ws.row_dimensions[3].height = 10
        
        # Headers
        headers = [
            "Test ID", "Component", "Description", "Action Steps", 
            "Expected Result", "Actual Result", "Status", 
            "Duration (s)", "Timestamp", "Screenshot Link"
        ]
        
        start_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        ws.row_dimensions[start_row].height = 25
        
        # Write Data
        current_row = start_row + 1
        for res in self.results:
            ws.cell(row=current_row, column=1, value=res["id"]).font = bold_data_font
            ws.cell(row=current_row, column=2, value=res["component"]).font = data_font
            ws.cell(row=current_row, column=3, value=res["description"]).font = data_font
            
            # Action Steps (allow multiline)
            steps_cell = ws.cell(row=current_row, column=4, value=res["steps"])
            steps_cell.font = data_font
            steps_cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            ws.cell(row=current_row, column=5, value=res["expected"]).font = data_font
            ws.cell(row=current_row, column=6, value=res["actual"]).font = data_font
            
            # Status styling
            status_cell = ws.cell(row=current_row, column=7, value=res["status"])
            status_cell.font = bold_data_font
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            if res["status"] == "PASS":
                status_cell.fill = pass_fill
            elif res["status"] in ("FAIL", "FAILED"):
                status_cell.fill = fail_fill
            else:
                status_cell.fill = error_fill
                
            duration_cell = ws.cell(row=current_row, column=8, value=res["time"])
            duration_cell.font = data_font
            duration_cell.alignment = Alignment(horizontal="right")
            
            ts_cell = ws.cell(row=current_row, column=9, value=res["timestamp"])
            ts_cell.font = data_font
            ts_cell.alignment = Alignment(horizontal="center")
            
            # Link to Screenshot if exists
            screenshot_cell = ws.cell(row=current_row, column=10)
            if res["screenshot"]:
                screenshot_cell.value = "View Screenshot"
                # Make it a clickable hyperlink relative to report file
                screenshot_cell.hyperlink = res["screenshot"]
                screenshot_cell.font = link_font
                screenshot_cell.alignment = Alignment(horizontal="center")
            else:
                screenshot_cell.value = "N/A"
                screenshot_cell.font = data_font
                screenshot_cell.alignment = Alignment(horizontal="center")
            
            # Apply borders and alignment for all data columns
            for col in range(1, 11):
                c = ws.cell(row=current_row, column=col)
                c.border = thin_border
                if col not in (4, 7, 8, 9, 10):  # Left align standard text, except steps/alignments
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    
            ws.row_dimensions[current_row].height = 60 # Set fixed comfortable height for wraps
            current_row += 1
            
        # Auto-fit columns with safety margins
        for col in ws.columns:
            # Skip merged title rows for width calculation
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 3 and cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        try:
            wb.save(str(REPORT_FILE_PATH))
            print(f"\n[TestReporter] Report generated successfully at: {REPORT_FILE_PATH}")
        except PermissionError:
            timestamp = int(time.time())
            alt_path = REPORT_FILE_PATH.parent / f"test_automation_report_{timestamp}.xlsx"
            wb.save(str(alt_path))
            print(f"\n[TestReporter] Default report file was locked. Saved instead at: {alt_path}")

def config_target_url():
    """Helper to fetch config app package safely without circular imports."""
    from config import APP_PACKAGE
    return APP_PACKAGE
